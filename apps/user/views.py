from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from openpyxl import load_workbook


from .models import UserProfile
from .serializer import UserProfileSerializer
from apps.user import serializer

class UserProfileView(APIView):
    def get(self, request):
        profiles = UserProfile.objects.all()
        serializer = UserProfileSerializer(profiles, many=True)
        if not serializer.data:
            return Response({"message": "No user profiles found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(serializer.data, status=status.HTTP_200_OK)


    def post(self, request):
        serializer = UserProfileSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request):
        pk = request.data.get('id')
        try:
            profile = UserProfile.objects.get(pk=pk)
            serializer = UserProfileSerializer(profile, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except UserProfile.DoesNotExist:
            return Response({"message": "User profile not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def delete(self, request):
        pk = request.data.get('id')
        try:
            profile = UserProfile.objects.get(pk=pk)
        except UserProfile.DoesNotExist:
            return Response({"message": "User profile not found."}, status=status.HTTP_404_NOT_FOUND)
        
        profile.delete()
        return Response(status=status.HTTP_204_NO_CONTENT, data={"message": "User profile deleted successfully."})
    


class SupportiveAuthView(APIView):
    def get(self, request):
        return Response({"message": "Supportive authentication endpoint is operational."}, status=status.HTTP_200_OK)
    
    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')

        profile = UserProfile.objects.filter(email=email, password=password).first()
        if profile:
            serializer = UserProfileSerializer(profile)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({"message": "Invalid email or password."}, status=status.HTTP_401_UNAUTHORIZED)




class BulkUserView(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"message": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = load_workbook(filename=file)
            sheet = wb.active

            created_users = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                user_name, email, password, name, is_active, is_admin, bio = row
                user_profile = UserProfile(
                    user_name=user_name,
                    email=email,
                    password=password,
                    name=name,
                    is_active=is_active,
                    is_admin=is_admin,
                    bio=bio
                )
                user_profile.save()
                created_users.append(user_profile)

            serializer = UserProfileSerializer(created_users, many=True)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"message": f"Error processing file: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)